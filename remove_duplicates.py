#! /usr/bin/python

import openpyxl

wb = openpyxl.load_workbook('Вычесть.xlsx')
sh = wb['Sheet1']

for i in range(1, sh.max_row):
    for j in range(1, sh.max_row):
        key = sh.cell(row=i, column=2)
        duplicate = sh.cell(row=j, column=1)
        if key.value is not None and duplicate.value is not None:
            if duplicate.value.lower() == key.value.lower():
                duplicate.value = None

wb.save('Вычесть.xlsx')
