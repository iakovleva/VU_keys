#! /usr/bin/python

import openpyxl
import enchant

d = enchant.Dict("ru-RU")

wb = openpyxl.load_workbook('bufer.xlsx')
sh = wb.get_sheet_by_name('Sheet1')

for i in range(1, sh.max_row):
    string = sh.cell(row=i, column=1).value
    if d.check(string):
        sh.cell(row=i, column=2).value = "True"

wb.save('bufer.xlsx')
