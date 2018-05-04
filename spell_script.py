#! /usr/bin/python

import openpyxl

wb = openpyxl.load_workbook('bufer.xlsx')
sh = wb.get_sheet_by_name('Sheet1')

for i in range(1, sh.max_row):
        key = sh.cell(row=i, column=6)
        for j in range(1, sh.max_row):
            duplicate = sh.cell(row=j, column=1)
            if i == 4148:
                print ("Finished")
                break
            else:
                if duplicate.value == key.value:
                    duplicate.value = None

wb.save('bufer.xlsx')
