#! /usr/bin/python
# -*- coding: utf-8 -*-

import sys
import openpyxl
import re

def insert_headline(fl):
    wb_data = openpyxl.load_workbook(fl)
    wb_example = openpyxl.load_workbook(filename)
    ws_data = wb_data.active
    ws = wb_example.active

    for i in range(1, ws_data.max_row+1):
        # Extract number after 'k'
        cell_value = ws_data.cell(row=i, column=7).value
        match = re.search(r'k\d+', cell_value)
        if match:
            num = int(match.group()[1:])
            headline_cell = ws_data.cell(row=i, column=2)
            description_cell = ws_data.cell(row=i, column=4)
            # Set headline
            if num < 1:
                print('>1, row ', i)
            elif num > 27:
               headline_cell.value = ws.cell(row=28, column=2).value
            else:
               headline_cell.value = ws.cell(row=num, column=2).value

            # Set description
            if num > 80:
                print('>80, row ', i)
            elif num < 31:
                description_cell.value = ws.cell(row=30, column=2).value
            else:
                description_cell.value = ws.cell(row=num, column=2).value

    wb_data.save(fl)

if __name__ == '__main__':
    insert_headline(sys.argv[1])
