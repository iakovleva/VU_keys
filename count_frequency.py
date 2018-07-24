#! /usr/bin/env python
# -*- coding: utf-8 -*-

import cgi
import sys, os
import openpyxl
from collections import Counter


HTML = """
<html>
<head>
<title></title>
</head>
<body>
  <h1>Upload File</h1>
  <form action="count_frequency.py" method="POST" enctype="multipart/form-data">
    File: <input name="file" type="file">
    <input name="submit" type="submit">
</form>

{% if filedata %}
<blockquote> {{filedata}} </blockquote>
{% endif %}
</body>
</html>
"""

form = cgi.FieldStorage()
UPLOAD_DIR='./uploads'
#fl = form.getvalue('fl')
uploaded_file = form['file']
fl = os.path.join(UPLOAD_DIR, os.path.basename(uploaded_file.filename))
col = form.getvalue('col')
common = form.getvalue('common')
all_data = form.getvalue('all_data')


def main(col, common, all_data):
    print("Content-type: text/html\r\n\r\n")
    print(HTML)
    print(fl)
    wb = openpyxl.load_workbook(fl)
    sh = wb.active
    sh2 = wb.create_sheet(title='Sheet1') # Create new sheet
    cnt = create_counter(sh, col)
    # Read arguments: common - number of most frequent words to display,
    # all_data - whether to write all columns or only words with occurences
#    for a in args:
#            common = int(args[0])
#            all_data = args[1]
    common = int(common)
    if all_data == 'True':
        copy_rows(sh, sh2, cnt, col, common)
    else:
        count_frequency(sh2, cnt, common)
    wb.save(fl)


def create_counter(sh, col):
    # Create Counter for all words in column
    cnt = Counter()
    for i in range(1, sh.max_row):
            cell = sh.cell(row=i, column=int(col)).value
            if cell is not None:
                for word in str(cell).split():
                    cnt[word.lower()] += 1
    print(len(cnt))
    return cnt

def count_frequency(sh2, cnt, common):
    row_num = 1
    # By default all common words are printed in Sheet2
    if common == 0:
        for k, v in cnt.items():
            sh2.cell(row=row_num, column=1).value = k
            sh2.cell(row=row_num, column=2).value = v
            row_num += 1
    # If 'common' number is set, this number of frequent words are printed
    else:
        for k, v in cnt.most_common(common):
            sh2.cell(row=row_num, column=1).value = k
            sh2.cell(row=row_num, column=2).value = v
            row_num += 1

def copy_rows(sh, sh2, cnt, col, common):
    # Create  headers row in new sheet
    sh2.cell(row=1, column=1).value = sh.cell(row=3, column=1).value
    sh2.cell(row=1, column=2).value = 'Word'
    for c in range(3, 18):
        sh2.cell(row=1, column=c).value = sh.cell(row=3, column=c-1).value

    row_num = 2

    # By default all common words are printed in Sheet2
    if common == 0:
        # Check if each common word is in each keyword phrase, create new row with all
        # data about kw plus a column with commom word in this phrase
        for k, v in cnt.items():
            for i in range(3, sh.max_row):
                cell = sh.cell(row=i, column=int(col)).value
                if cell is not None:
                    if k in str(cell).lower().split():
                        sh2.cell(row=row_num, column=1).value = sh.cell(row=i, column=1).value
                        sh2.cell(row=row_num, column=2).value = k
                        sh2.cell(row=row_num, column=3).value = sh.cell(row=i, column=2).value
                        sh2.cell(row=row_num, column=4).value = sh.cell(row=i, column=3).value
                        sh2.cell(row=row_num, column=5).value = sh.cell(row=i, column=4).value
                        sh2.cell(row=row_num, column=6).value = sh.cell(row=i, column=5).value
                        sh2.cell(row=row_num, column=7).value = sh.cell(row=i, column=6).value
                        sh2.cell(row=row_num, column=8).value = sh.cell(row=i, column=7).value
                        sh2.cell(row=row_num, column=9).value = sh.cell(row=i, column=8).value
                        sh2.cell(row=row_num, column=10).value = sh.cell(row=i, column=9).value
                        sh2.cell(row=row_num, column=11).value = sh.cell(row=i, column=10).value
                        sh2.cell(row=row_num, column=12).value = sh.cell(row=i, column=11).value
                        sh2.cell(row=row_num, column=13).value = sh.cell(row=i, column=12).value
                        sh2.cell(row=row_num, column=14).value = sh.cell(row=i, column=13).value
                        sh2.cell(row=row_num, column=15).value = sh.cell(row=i, column=14).value
                        sh2.cell(row=row_num, column=16).value = sh.cell(row=i, column=15).value
                        sh2.cell(row=row_num, column=17).value = sh.cell(row=i, column=16).value
                        row_num += 1

    # If 'common' number is set, this number of frequent words are printed
    else:
        for k, v in cnt.most_common(common):
            for i in range(3, sh.max_row):
                cell = sh.cell(row=i, column=int(col)).value
                if cell is not None:
                    if k in str(cell).lower().split():
                        sh2.cell(row=row_num, column=1).value = sh.cell(row=i, column=1).value
                        sh2.cell(row=row_num, column=2).value = k
                        sh2.cell(row=row_num, column=3).value = sh.cell(row=i, column=2).value
                        sh2.cell(row=row_num, column=4).value = sh.cell(row=i, column=3).value
                        sh2.cell(row=row_num, column=5).value = sh.cell(row=i, column=4).value
                        sh2.cell(row=row_num, column=6).value = sh.cell(row=i, column=5).value
                        sh2.cell(row=row_num, column=7).value = sh.cell(row=i, column=6).value
                        sh2.cell(row=row_num, column=8).value = sh.cell(row=i, column=7).value
                        sh2.cell(row=row_num, column=9).value = sh.cell(row=i, column=8).value
                        sh2.cell(row=row_num, column=10).value = sh.cell(row=i, column=9).value
                        sh2.cell(row=row_num, column=11).value = sh.cell(row=i, column=10).value
                        sh2.cell(row=row_num, column=12).value = sh.cell(row=i, column=11).value
                        sh2.cell(row=row_num, column=13).value = sh.cell(row=i, column=12).value
                        sh2.cell(row=row_num, column=14).value = sh.cell(row=i, column=13).value
                        sh2.cell(row=row_num, column=15).value = sh.cell(row=i, column=14).value
                        sh2.cell(row=row_num, column=16).value = sh.cell(row=i, column=15).value
                        sh2.cell(row=row_num, column=17).value = sh.cell(row=i, column=16).value
                        row_num += 1


if __name__ == '__main__':
#    main(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
    main(col, common, all_data)
